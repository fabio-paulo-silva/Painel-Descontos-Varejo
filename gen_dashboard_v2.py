#!/usr/bin/env python3
# gen_dashboard_v2.py — Dashboard Auditoria de Descontos
import csv, json, re, unicodedata
from collections import defaultdict
from datetime import datetime

def norm_name(s):
    s = (s or '').strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    return ' '.join(s.split())

import os
BASE_DIR = r'C:\Users\fabio.silva\OneDrive - Gentil Negócios\Área de Trabalho\Fábio\merge_bases_descontos'
DIST = os.path.join(BASE_DIR, 'dist')           # pasta publicada no GitHub Pages
os.makedirs(DIST, exist_ok=True)
CSV_PATH = os.path.join(BASE_DIR, 'merge_consolidado_geral.csv')
OUT_PATH = os.path.join(DIST, 'index.html')      # dashboard (página inicial do Pages)
DCENTROS_PATH = os.path.join(BASE_DIR, 'dCentros.xlsx')

def norm_code(v):
    if v is None:
        return ''
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s

# ── Carrega dCentros (cadastro de lojas): BCPS -> nome/praça/regional/cluster ──
# Também coleta os nomes dos GESTORES (GVO, GCVO/GPVO, GRVO) para excluí-los da
# análise por consultor (vendas no nome do gestor distorcem o % por consultor).
DCENTROS = {}
GESTORES = set()  # nomes de gestores normalizados
def carregar_dcentros():
    try:
        import openpyxl
    except Exception as e:
        print("AVISO: openpyxl indisponível, dCentros ignorado:", e)
        return
    try:
        wb = openpyxl.load_workbook(DCENTROS_PATH, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        hdr = next(it)
        idx = {str(h).strip().upper(): i for i, h in enumerate(hdr) if h is not None}
        ci = idx.get('BCPS'); ni = idx.get('LOJA'); ri = idx.get('REGIONAL'); cli = idx.get('CLUSTER')
        pi = next((idx[k] for k in idx if k.startswith('PRA')), None)
        gest_cols = [idx[k] for k in ('GVO', 'GCVO/GPVO', 'GRVO') if k in idx]
        for r in it:
            if r is None:
                continue
            code = norm_code(r[ci]) if ci is not None else ''
            if not code:
                continue
            def cell(i):
                return str(r[i]).strip() if (i is not None and i < len(r) and r[i] is not None) else ''
            DCENTROS[code] = {'nome': cell(ni), 'praca': cell(pi), 'regional': cell(ri), 'cluster': cell(cli)}
            for gi in gest_cols:
                g = norm_name(cell(gi))
                if g and len(g.split()) >= 2:  # ignora vazios/genéricos
                    GESTORES.add(g)
        print(f"dCentros: {len(DCENTROS)} lojas carregadas | {len(GESTORES)} gestores")
    except Exception as e:
        print("AVISO: falha ao carregar dCentros:", e)

carregar_dcentros()

# Gestoras cujo nome na base de vendas não casa com o cadastro dCentros (informado
# manualmente). Comparado contra o nome CURTO (primeiro+último) do consultor:
# "ANDREA COSTA NOGUEIRA MELONIO" -> "ANDREA MELONIO"; "THAIS ELLEN CARTAGENES SOARES" -> "THAIS SOARES".
EXTRA_GESTOR_SHORT = {'THAIS SOARES', 'ANDREA MELONIO'}

_gestor_cache = {}
def is_gestor(full_name):
    n = norm_name(full_name)
    if not n:
        return False
    if n in _gestor_cache:
        return _gestor_cache[n]
    res = False
    if n in GESTORES:
        res = True
    else:
        for g in GESTORES:
            if n == g or n.startswith(g + ' ') or g.startswith(n + ' '):
                res = True
                break
    _gestor_cache[n] = res
    return res

def loja_info(code, fallback_name):
    d = DCENTROS.get(code)
    if d and d['nome']:
        return (d['nome'],
                d['regional'] or '(sem regional)',
                d['praca'] or '(sem praça)',
                d['cluster'] or '(sem cluster)')
    return (fallback_name, '(sem regional)', '(sem praça)', '(sem cluster)')

def parse_br(s):
    if not s or s.strip() == '':
        return 0.0
    return float(s.strip().replace('.', '').replace(',', '.'))

def to_iso(data_str):
    # DD/MM/YYYY -> YYYY-MM-DD (para filtro de data e ordenação)
    s = (data_str or '').strip()
    if '/' in s:
        p = s.split('/')
        if len(p) == 3:
            return f"{p[2]}-{p[1].zfill(2)}-{p[0].zfill(2)}"
    return s

def mes_label(mes_key):
    m = {'2026-01':'Jan/2026','2026-02':'Fev/2026','2026-03':'Mar/2026',
         '2026-04':'Abr/2026','2026-05':'Mai/2026','2026-06':'Jun/2026',
         '2026-07':'Jul/2026','2026-08':'Ago/2026','2026-09':'Set/2026',
         '2026-10':'Out/2026','2026-11':'Nov/2026','2026-12':'Dez/2026',
         '2025-01':'Jan/2025','2025-02':'Fev/2025','2025-03':'Mar/2025',
         '2025-04':'Abr/2025','2025-05':'Mai/2025','2025-06':'Jun/2025',
         '2025-07':'Jul/2025','2025-08':'Ago/2025','2025-09':'Set/2025',
         '2025-10':'Out/2025','2025-11':'Nov/2025','2025-12':'Dez/2025'}
    return m.get(mes_key, mes_key)

def short_name(full):
    parts = full.strip().split()
    if len(parts) <= 2:
        return full.strip()
    return parts[0] + ' ' + parts[-1]

def trunc(s, n):
    s = s.strip()
    return s[:n] if len(s) > n else s

print("Lendo CSV...")

# First pass: aggregate coupon-SKU records
# key = (chave_unica, sku)
sku_data = {}  # key -> {bruto, promo, manual, fidelidade, loja, consultor, canal, mes, motivo2, campanha, hora, data_raw}
gestor_short = set()  # nomes (short) de consultores que são gestores -> excluídos da análise por consultor

VALID_ORIGINS = {'PROMOCIONAL', 'MANUAL', 'FIDELIDADE', 'SEM DESCONTOS'}

with open(CSV_PATH, encoding='utf-8-sig', newline='') as f:
    reader = csv.reader(f, delimiter=';')
    header = next(reader)
    # Map column names to indices
    col = {h.strip(): i for i, h in enumerate(header)}

    # Column indices
    i_codloja = col.get('DESC_Codigo Loja', 0)
    i_loja = col.get('DESC_Loja', 1)
    i_origem = col.get('DESC_Origem Desconto', 2)
    i_campanha = col.get('DESC_Descricao Campanha', 3)
    i_data = col.get('DESC_Data Desconto', 5)
    i_sku = col.get('DESC_Codigo SKU', 7)
    i_produto = col.get('DESC_Descricao Produto', 8)
    i_qtd = col.get('DESC_Qtd.', 9)
    i_bruto = col.get('DESC_Valor Bruto', 10)
    i_desc = col.get('DESC_Valor Desconto', 11)
    i_liquido = col.get('DESC_Valor Liquido', 12)
    i_chave = col.get('Chave Unica', 14)
    i_hora = col.get('CUP_Hora', 15)
    i_canal = col.get('CUP_Canal de venda', 16)
    i_motivo2 = col.get('Motivo Desconto 2', 23)
    i_consultor = col.get('Nome Consultor', 26)

    rows_read = 0
    for row in reader:
        rows_read += 1
        if rows_read % 200000 == 0:
            print(f"  {rows_read:,} linhas...")
        if len(row) < 20:
            continue

        chave = row[i_chave].strip()
        sku = row[i_sku].strip()
        origem = row[i_origem].strip().upper()
        if not chave or not sku:
            continue

        key = (chave, sku)
        desc_val = parse_br(row[i_desc])

        if key not in sku_data:
            # Parse date -> mes
            data_str = row[i_data].strip()
            mes_key = ''
            try:
                if '/' in data_str:
                    parts = data_str.split('/')
                    mes_key = f"{parts[2]}-{parts[1]}"
                elif '-' in data_str:
                    parts = data_str.split('-')
                    mes_key = f"{parts[0]}-{parts[1]}"
            except:
                pass

            loja_raw = row[i_loja].strip()
            consultor_raw = row[i_consultor].strip()
            consultor_short = trunc(short_name(consultor_raw), 35)
            if is_gestor(consultor_raw) or norm_name(consultor_short) in EXTRA_GESTOR_SHORT:
                gestor_short.add(consultor_short)
            canal_raw = row[i_canal].strip()
            motivo2_raw = row[i_motivo2].strip()
            campanha_raw = row[i_campanha].strip()
            hora_raw = row[i_hora].strip()

            sku_data[key] = {
                'loja': trunc(loja_raw, 40),
                'consultor': consultor_short,
                'canal': trunc(canal_raw, 25),
                'mes': mes_key,
                'motivo2': trunc(motivo2_raw, 40),
                'campanha': trunc(campanha_raw, 50),
                'hora': hora_raw[:5] if hora_raw else '',
                'data': data_str,
                'codloja': norm_code(row[i_codloja]),
                'sku': sku,
                'produto': trunc(row[i_produto], 45),
                # O bruto se REPETE em cada linha de desconto do mesmo SKU (descontos
                # empilhados). Logo: bruto = MAIOR bruto de uma linha (contado 1x);
                # descontos = SOMA por origem. Duplicidade real = linha (origem+desc)
                # idêntica repetida -> 'sig' conta repetições por assinatura.
                'mb': 0.0,   # maior bruto de uma linha
                'mq': 0.0,   # maior qtd de uma linha
                'p': 0.0, 'm': 0.0, 'f': 0.0,  # somas de desconto por origem
                'sig': {},   # (origem,desc) -> contagem (p/ detectar repetição idêntica)
            }

        rec = sku_data[key]
        bval = parse_br(row[i_bruto])
        qval = parse_br(row[i_qtd])
        if bval > rec['mb']:
            rec['mb'] = bval
        if qval > rec['mq']:
            rec['mq'] = qval
        if origem == 'PROMOCIONAL':
            rec['p'] += desc_val; oc = 'P'
        elif origem == 'MANUAL':
            rec['m'] += desc_val; oc = 'M'
        elif origem == 'FIDELIDADE':
            rec['f'] += desc_val; oc = 'F'
        else:
            oc = 'S'
        sg = (oc, round(desc_val, 2))
        rec['sig'][sg] = rec['sig'].get(sg, 0) + 1

print(f"Total coupon-SKU únicos: {len(sku_data):,}")

# Finalização por SKU — corrige bruto (múltiplos escaneamentos do mesmo SKU+origem
# inflavam o desconto). Bruto real = maior soma de bruto entre as origens (cada origem
# repete o bruto do item; somar entre origens multiplicaria). Detecta duplicidade.
print("Finalizando SKUs (correção de duplicidade)...")
item_acc = defaultdict(lambda: [0.0, 0.0, 0.0])  # (mes, sku) -> [bruto, promo, manual]
item_name = {}
dup_total = 0
for rec in sku_data.values():
    bruto = rec['mb']
    promo = rec['p']; manual = rec['m']; fidelidade = rec['f']
    qtd = rec['mq']
    scans = max(rec['sig'].values()) if rec['sig'] else 1  # maior repetição idêntica
    rec['bruto'] = bruto
    rec['promo'] = promo
    rec['manual'] = manual
    rec['fidelidade'] = fidelidade
    rec['qtd'] = qtd
    rec['scans'] = scans
    rec['dup'] = 1 if scans > 1 else 0
    if rec['dup']:
        dup_total += 1
    del rec['mb'], rec['mq'], rec['p'], rec['m'], rec['f'], rec['sig']
    # acumula análise por item (rede), separável por mês
    k = (rec['mes'], rec['sku'])
    ia = item_acc[k]
    ia[0] += bruto; ia[1] += promo; ia[2] += manual
    if rec['sku'] not in item_name:
        item_name[rec['sku']] = rec['produto']

print(f"SKUs com escaneamento duplicado (qtd digitada manualmente repetida): {dup_total:,}")

# Monta arrays de itens para os gráficos de % por item
items_prod_idx, ITEMS_PROD, ITEMS = {}, [], []
for (mes, sku), ia in item_acc.items():
    nm = item_name.get(sku, sku)
    if nm not in items_prod_idx:
        items_prod_idx[nm] = len(ITEMS_PROD); ITEMS_PROD.append(nm)
    ITEMS.append([items_prod_idx[nm], mes, round(ia[0], 2), round(ia[1], 2), round(ia[2], 2)])
print(f"Itens (mes×sku) para análise de %: {len(ITEMS):,}")

# Second pass: aggregate to (mes, loja, consultor, canal, motivo2, campanha)
# ROWS header: mes, loja, consultor, canal, motivo2, campanha, bruto, promo, manual, fidelidade, cnt, viol_a, viol_c, f5_cnt
group_data = defaultdict(lambda: [0.0, 0.0, 0.0, 0.0, 0, 0, 0, 0, 0, 0])
# indices:                                                bruto promo manual fid cnt va vc f5 f60 dup

violations = []      # top violations (A/C) for table
exceptions_f = []    # Regra F: desconto total > 60% do bruto (qualquer origem)

print("Agregando grupos...")

cod_meta = {}  # codloja -> (nome, regional, praca, cluster)
for (chave, sku), rec in sku_data.items():
    mes = rec['mes']
    data_iso = to_iso(rec['data'])
    codloja = rec.get('codloja', '')
    if codloja not in cod_meta:
        cod_meta[codloja] = loja_info(codloja, rec['loja'])
    loja_nome, regional, praca, cluster = cod_meta[codloja]
    loja = loja_nome
    consultor = rec['consultor']
    canal = rec['canal']
    motivo2 = rec['motivo2']
    campanha = rec['campanha']
    bruto = rec['bruto']
    promo = rec['promo']
    manual = rec['manual']
    fidelidade = rec['fidelidade']

    # Business rules
    viol_a = 1 if ('11 -' in canal and manual > 0) else 0
    viol_c = 1 if (manual > 0 and (not motivo2 or motivo2 == '0 - Não é desconto fidelidade')) else 0
    liquido_real = bruto - promo - manual - fidelidade
    f5 = 1 if liquido_real <= 0 else 0
    # Regra F — alerta-teto absoluto: desconto total (todas as origens) > 60% do bruto
    desc_total = promo + manual + fidelidade
    pct_desc = (desc_total / bruto) if bruto > 0 else 0.0
    viol_f = 1 if (bruto > 0 and pct_desc > 0.60) else 0

    grp_key = (data_iso, codloja, consultor, canal, motivo2, campanha)
    g = group_data[grp_key]
    g[0] += bruto
    g[1] += promo
    g[2] += manual
    g[3] += fidelidade
    g[4] += 1
    g[5] += viol_a
    g[6] += viol_c
    g[7] += f5
    g[8] += viol_f
    g[9] += rec['dup']

    # Collect violations
    if (viol_a or viol_c) and manual > 0:
        regra = []
        if viol_a:
            regra.append('A')
        if viol_c:
            regra.append('C')
        violations.append({
            'loja': loja,
            'consultor': consultor,
            'canal': canal,
            'data': data_iso,
            'regional': regional,
            'praca': praca,
            'cluster': cluster,
            'manual': round(manual, 2),
            'bruto': round(bruto, 2),
            'motivo2': motivo2,
            'regra': '+'.join(regra)
        })

    # Collect Regra F exceptions (any origin, > 60% do bruto)
    if viol_f:
        if promo >= manual and promo >= fidelidade:
            origem = 'Promocional'
        elif manual >= fidelidade:
            origem = 'Manual'
        else:
            origem = 'Fidelidade'
        exceptions_f.append({
            'loja': loja,
            'consultor': consultor,
            'canal': canal,
            'data': data_iso,
            'regional': regional,
            'praca': praca,
            'cluster': cluster,
            'bruto': round(bruto, 2),
            'desc_total': round(desc_total, 2),
            'pct': round(pct_desc * 100, 1),
            'origem': origem,
        })

# Sort violations by manual desc, keep top 200
violations.sort(key=lambda x: -x['manual'])
violations = violations[:200]

# Sort Regra F exceptions by valor exposto (desc_total) desc, keep top 200
exceptions_f.sort(key=lambda x: -x['desc_total'])
exceptions_f = exceptions_f[:200]

# ─────────────────────────────────────────────────────────────────────────────
# Detalhamento cupom a cupom (arquivo companheiro, carregado sob demanda)
# ─────────────────────────────────────────────────────────────────────────────
print("Construindo detalhamento por cupom...")

OUT_DETALHE = r'C:\Users\fabio.silva\OneDrive - Gentil Negócios\Área de Trabalho\Fábio\merge_bases_descontos\detalhamento_dados.js'

# Lookup tables (dedupe strings)
lojas_idx, cons_idx, prod_idx, datas_idx, mot_idx, camp_idx = {}, {}, {}, {}, {}, {}
lojas_list, cons_list, prod_list, datas_list, mot_list, camp_list = [], [], [], [], [], []
def _idx(val, d, lst):
    if val not in d:
        d[val] = len(lst); lst.append(val)
    return d[val]

# lojas do detalhamento mantêm índice por CÓDIGO, com dimensões paralelas
lojas_reg, lojas_praca, lojas_cluster = [], [], []
def _idx_loja_det(code):
    if code not in lojas_idx:
        lojas_idx[code] = len(lojas_list)
        nome, reg, praca, cluster = cod_meta.get(code) or loja_info(code, code)
        lojas_list.append(nome); lojas_reg.append(reg); lojas_praca.append(praca); lojas_cluster.append(cluster)
    return lojas_idx[code]

# Group coupon-SKU records by Chave Unica
coupons = {}  # chave -> {cod,data,cons,mes, b,p,m,f, items:[]}
for (chave, sku), rec in sku_data.items():
    cup = coupons.get(chave)
    if cup is None:
        cup = {'cod': rec.get('codloja', ''), 'data': rec['data'], 'cons': rec['consultor'],
               'mes': rec['mes'], 'b': 0.0, 'p': 0.0, 'm': 0.0, 'f': 0.0, 'items': []}
        coupons[chave] = cup
    cup['b'] += rec['bruto']; cup['p'] += rec['promo']
    cup['m'] += rec['manual']; cup['f'] += rec['fidelidade']
    cup['items'].append((rec.get('sku', ''), rec.get('produto', ''), rec.get('qtd', 0.0),
                         rec['bruto'], rec['promo'], rec['manual'], rec['fidelidade'],
                         rec.get('scans', 1), rec.get('motivo2', ''), rec.get('campanha', '')))

# Build compact CUP array
CUP = []
for chave, cup in coupons.items():
    li = _idx_loja_det(cup['cod'])
    di = _idx(cup['data'], datas_idx, datas_list)
    ci = _idx(cup['cons'], cons_idx, cons_list)
    boleto = ''
    try:
        boleto = chave.split('-')[1]
    except Exception:
        boleto = chave
    items = []
    for (sk, pr, q, b, p, m, f, sc, mot, camp) in cup['items']:
        pi = _idx(pr, prod_idx, prod_list)
        mi = _idx(mot, mot_idx, mot_list)
        cmi = _idx(camp, camp_idx, camp_list)
        items.append([sk, pi, round(q, 0), round(b, 2), round(p, 2), round(m, 2), round(f, 2), sc, mi, cmi])
    CUP.append([
        boleto, li, di, ci, cup['mes'],
        round(cup['b'], 2), round(cup['p'], 2), round(cup['m'], 2), round(cup['f'], 2),
        items
    ])

# Sort coupons by total desconto desc (mais relevantes primeiro)
CUP.sort(key=lambda c: -(c[6] + c[7] + c[8]))

# Particiona por MÊS (escala o ano inteiro sem estourar 100MB/arquivo nem travar o navegador)
cup_by_month = defaultdict(list)
for c in CUP:
    cup_by_month[c[4] or 'sem-data'].append(c)
meses_det = sorted(cup_by_month.keys())

# Tabela base (lookups compartilhados) — carregada 1x ao abrir a aba
det_base = {
    'lojas': lojas_list, 'cons': cons_list, 'prod': prod_list, 'datas': datas_list,
    'loja_reg': lojas_reg, 'loja_praca': lojas_praca, 'loja_cluster': lojas_cluster,
    'regionais': sorted(set(lojas_reg)), 'pracas': sorted(set(lojas_praca)), 'clusters': sorted(set(lojas_cluster)),
    'mot': mot_list, 'camp': camp_list,
    'meses': meses_det,
    'mes_labels': {k: mes_label(k) for k in meses_det},
}
with open(os.path.join(DIST, 'detalhe_base.js'), 'w', encoding='utf-8') as f:
    f.write('window.DET_BASE=')
    json.dump(det_base, f, ensure_ascii=False, separators=(',', ':'))
    f.write(';')

# Um arquivo de cupons por mês (carregado sob demanda)
for m, cups in cup_by_month.items():
    fn = os.path.join(DIST, f'detalhe_{m}.js')
    with open(fn, 'w', encoding='utf-8') as f:
        f.write(f"window.DET_M=window.DET_M||{{}};window.DET_M[{json.dumps(m)}]=")
        json.dump(cups, f, ensure_ascii=False, separators=(',', ':'))
        f.write(';')
    print(f"  detalhe_{m}.js: {len(cups):,} cupons ({os.path.getsize(fn)/1024/1024:.1f} MB)")
print(f"Detalhamento particionado em {len(cup_by_month)} mes(es) na pasta dist/")

# Build ROWS array (strings normalizadas em índices -> arquivo leve)
# Colunas: mes, data(ISO), lojaIdx, consIdx, canalIdx, motivoIdx, campIdx, bruto, promo,
#          manual, fidelidade, cnt, viol_a, viol_c, f5_cnt, f60_cnt, dup_cnt
ROWS_HEADER = ["mes","data","loja","consultor","canal","motivo2","campanha",
               "bruto","promo","manual","fidelidade","cnt","viol_a","viol_c","f5_cnt","f60_cnt","dup_cnt"]
L_loja, L_cons, L_canal, L_mot, L_camp = {}, {}, {}, {}, {}
A_loja, A_cons, A_canal, A_mot, A_camp = [], [], [], [], []
A_loja_reg, A_loja_praca, A_loja_cluster = [], [], []
def _ix(v, d, a):
    if v not in d:
        d[v] = len(a); a.append(v)
    return d[v]
def _ix_loja(code):
    if code not in L_loja:
        L_loja[code] = len(A_loja)
        nome, reg, praca, cluster = cod_meta.get(code) or loja_info(code, code)
        A_loja.append(nome); A_loja_reg.append(reg); A_loja_praca.append(praca); A_loja_cluster.append(cluster)
    return L_loja[code]

rows_list = []
for (data_iso, codloja, consultor, canal, motivo2, campanha), g in group_data.items():
    rows_list.append([
        data_iso[:7], data_iso,
        _ix_loja(codloja), _ix(consultor, L_cons, A_cons),
        _ix(canal, L_canal, A_canal), _ix(motivo2, L_mot, A_mot), _ix(campanha, L_camp, A_camp),
        round(g[0], 2), round(g[1], 2), round(g[2], 2), round(g[3], 2),
        g[4], g[5], g[6], g[7], g[8], g[9]
    ])

print(f"Grupos ROWS (por data): {len(rows_list):,}")
print(f"Violações coletadas: {len(violations):,}")

regionais = sorted(set(A_loja_reg))
pracas = sorted(set(A_loja_praca))
clusters = sorted(set(A_loja_cluster))
cons_gestor = [1 if nm in gestor_short else 0 for nm in A_cons]
print(f"Consultores marcados como gestores (excluídos da análise por consultor): {sum(cons_gestor)}")
lookups_json = json.dumps({
    'loja': A_loja, 'cons': A_cons, 'canal': A_canal, 'motivo2': A_mot, 'campanha': A_camp,
    'loja_reg': A_loja_reg, 'loja_praca': A_loja_praca, 'loja_cluster': A_loja_cluster,
    'regionais': regionais, 'pracas': pracas, 'clusters': clusters,
    'cons_gestor': cons_gestor
}, ensure_ascii=False, separators=(',', ':'))

# Generate date info
gen_date = datetime.now().strftime('%d/%m/%Y %H:%M')

# Compute mes range for subtitle
meses_presentes = sorted(set(r[0] for r in rows_list if r[0]))
periodo = ' a '.join([mes_label(m) for m in [meses_presentes[0], meses_presentes[-1]]]) if len(meses_presentes) >= 2 else (mes_label(meses_presentes[0]) if meses_presentes else '')

# Dados principais (agregados) num arquivo externo -> HTML fica leve e cacheável
mes_labels_obj = {k: mes_label(k) for k in meses_presentes}
with open(os.path.join(DIST, 'painel_dados.js'), 'w', encoding='utf-8') as f:
    f.write('window.PAINEL=')
    json.dump({
        'ROWS_HEADER': ROWS_HEADER,
        'ROWS': rows_list,
        'VIOLATIONS': violations,
        'EXCEPTIONS_F': exceptions_f,
        'ITEMS': ITEMS,
        'ITEMS_PROD': ITEMS_PROD,
        'MES_LABELS': mes_labels_obj,
        'LK': json.loads(lookups_json),
    }, f, ensure_ascii=False, separators=(',', ':'))
    f.write(';')
print(f"painel_dados.js: {os.path.getsize(os.path.join(DIST, 'painel_dados.js'))/1024/1024:.1f} MB")

print("Gerando HTML...")

HTML = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Dashboard Auditoria de Descontos</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root {{
  --bg: #000000;
  --card: #1C2B3A;
  --border: #2a3f54;
  --text: #FFFFFF;
  --sub: #8aa3b8;
  --green: #2ECC8A;
  --blue: #3355FF;
  --cyan: #00AAFF;
  --mint: #4DE8AA;
  --red: #FF4455;
  --yellow: #FFB830;
  --navy: #1C2B3A;
}}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ background: var(--bg); color: var(--text); font-family: Arial, sans-serif; font-size: 14px; }}
a {{ color: var(--cyan); }}

/* HEADER */
.header {{
  background: var(--navy);
  border-bottom: 3px solid var(--green);
  padding: 18px 32px;
  display: flex;
  align-items: center;
  gap: 20px;
}}
.header-logo {{
  width: 48px; height: 48px;
  background: var(--green);
  border-radius: 10px;
  display: flex; align-items: center; justify-content: center;
  font-size: 24px; font-weight: bold; color: #000;
  flex-shrink: 0;
}}
.header-text h1 {{ font-size: 20px; font-weight: 700; color: var(--text); }}
.header-text p {{ font-size: 12px; color: var(--sub); margin-top: 3px; }}

/* TABS */
.tabs {{
  background: var(--card);
  border-bottom: 1px solid var(--border);
  display: flex;
  padding: 0 32px;
  gap: 4px;
}}
.tab-btn {{
  background: none; border: none; border-bottom: 3px solid transparent;
  color: var(--sub); font-size: 14px; font-weight: 600;
  padding: 14px 20px; cursor: pointer;
  transition: color 0.2s, border-color 0.2s;
  white-space: nowrap;
}}
.tab-btn:hover {{ color: var(--text); }}
.tab-btn.active {{ color: var(--green); border-bottom-color: var(--green); }}

/* FILTERS */
.filters-bar {{
  background: var(--card);
  border-bottom: 1px solid var(--border);
  padding: 12px 32px;
  display: flex;
  flex-wrap: wrap;
  gap: 10px;
  align-items: center;
}}
.filter-group {{ display: flex; flex-direction: column; gap: 3px; min-width: 130px; }}
.filter-group label {{ font-size: 11px; color: var(--sub); text-transform: uppercase; letter-spacing: 0.5px; }}
.filter-group select {{
  background: #0d1f2d; border: 1px solid var(--border); color: var(--text);
  border-radius: 6px; padding: 6px 10px; font-size: 13px; cursor: pointer;
  outline: none; min-width: 130px;
}}
.filter-group select:focus {{ border-color: var(--green); }}
.filter-group input {{
  background: #0d1f2d; border: 1px solid var(--border); color: var(--text);
  border-radius: 6px; padding: 6px 10px; font-size: 13px; min-width: 150px; outline: none;
}}
.filter-group input:focus {{ border-color: var(--green); }}
.filter-group input[type=date] {{ min-width: 140px; color-scheme: dark; }}
.filter-group input::placeholder {{ color: #5a7a92; }}
.btn-clear {{
  background: transparent; border: 1px solid var(--red); color: var(--red);
  border-radius: 6px; padding: 7px 16px; cursor: pointer; font-size: 13px;
  font-weight: 600; align-self: flex-end; transition: background 0.2s;
}}
.btn-clear:hover {{ background: var(--red); color: #fff; }}

/* CONTENT */
.content {{ padding: 24px 32px; }}
.page {{ display: none; }}
.page.active {{ display: block; }}

/* KPI CARDS */
.kpi-row {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin-bottom: 24px; }}
.kpi-card {{
  background: var(--card); border-radius: 10px; padding: 18px 20px;
  border-left: 4px solid var(--green);
  position: relative; overflow: hidden;
}}
.kpi-card.red {{ border-left-color: var(--red); }}
.kpi-card.yellow {{ border-left-color: var(--yellow); }}
.kpi-card.blue {{ border-left-color: var(--blue); }}
.kpi-card.cyan {{ border-left-color: var(--cyan); }}
.kpi-label {{ font-size: 11px; color: var(--sub); text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 8px; }}
.kpi-value {{ font-size: 26px; font-weight: 700; color: var(--text); line-height: 1; }}
.kpi-sub {{ font-size: 12px; color: var(--sub); margin-top: 6px; }}

/* CHARTS */
.charts-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 24px; }}
.charts-grid.cols-3 {{ grid-template-columns: 1fr 1fr 1fr; }}
.chart-card {{
  background: var(--card); border-radius: 10px; padding: 18px 20px;
  border: 1px solid var(--border);
}}
.chart-card.full {{ grid-column: 1 / -1; }}
.chart-title {{ font-size: 13px; font-weight: 600; color: var(--sub); margin-bottom: 14px; text-transform: uppercase; letter-spacing: 0.5px; }}
.chart-wrap {{ position: relative; }}
.chart-wrap canvas {{ max-height: 280px; }}
.chart-wrap.tall canvas {{ max-height: 380px; }}
.chart-scroll {{ height: 360px; overflow-y: auto; overflow-x: hidden; border-radius: 6px; }}
.chart-scroll::-webkit-scrollbar {{ width: 9px; }}
.chart-scroll::-webkit-scrollbar-thumb {{ background: #33506a; border-radius: 5px; }}
.chart-holder {{ position: relative; width: 100%; }}

/* TABLES */
.table-section {{ margin-top: 8px; }}
.table-title {{ font-size: 15px; font-weight: 700; color: var(--text); margin-bottom: 12px; }}
.table-wrap {{ overflow-x: auto; border-radius: 10px; border: 1px solid var(--border); }}
table {{ width: 100%; border-collapse: collapse; background: var(--card); }}
thead tr {{ background: #0d1f2d; }}
th {{ padding: 10px 14px; font-size: 11px; color: var(--sub); text-align: left; text-transform: uppercase; letter-spacing: 0.5px; border-bottom: 1px solid var(--border); white-space: nowrap; }}
td {{ padding: 9px 14px; font-size: 13px; color: var(--text); border-bottom: 1px solid var(--border); }}
tbody tr:hover {{ background: rgba(255,255,255,0.03); }}
tbody tr:last-child td {{ border-bottom: none; }}
.tag {{ display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 700; }}
.tag-red {{ background: rgba(255,68,85,0.2); color: var(--red); }}
.tag-yellow {{ background: rgba(255,184,48,0.2); color: var(--yellow); }}
.tag-green {{ background: rgba(46,204,138,0.2); color: var(--green); }}

/* PAGINATION */
.pagination {{ display: flex; gap: 8px; align-items: center; margin-top: 12px; justify-content: flex-end; }}
.pagination button {{
  background: var(--card); border: 1px solid var(--border); color: var(--text);
  padding: 5px 12px; border-radius: 5px; cursor: pointer; font-size: 12px;
}}
.pagination button:hover {{ border-color: var(--green); }}
.pagination button.active {{ background: var(--green); color: #000; border-color: var(--green); }}
.pagination span {{ color: var(--sub); font-size: 12px; }}

.no-data {{ color: var(--sub); text-align: center; padding: 40px; font-size: 14px; }}
.legenda {{ background: rgba(46,204,138,0.08); border: 1px solid var(--border); border-left: 3px solid var(--green); border-radius: 6px; padding: 8px 12px; margin-bottom: 10px; font-size: 12px; color: var(--sub); line-height: 1.5; }}
.legenda b {{ color: var(--text); }}

/* DETALHAMENTO */
.det-filters {{ display: flex; flex-wrap: wrap; gap: 12px; align-items: flex-end; background: var(--card); border: 1px solid var(--border); border-radius: 10px; padding: 16px 18px; margin-bottom: 18px; }}
.det-filters input {{ background: #0d1f2d; border: 1px solid var(--border); color: var(--text); border-radius: 6px; padding: 6px 10px; font-size: 13px; min-width: 150px; outline: none; }}
.det-filters input:focus {{ border-color: var(--green); }}
.btn-search {{ background: var(--green); border: none; color: #000; font-weight: 700; border-radius: 6px; padding: 8px 20px; cursor: pointer; font-size: 13px; align-self: flex-end; }}
.btn-search:hover {{ background: #25b377; }}
.det-summary {{ display: flex; flex-wrap: wrap; gap: 14px; margin-bottom: 16px; }}
.det-summary .box {{ background: var(--card); border: 1px solid var(--border); border-left: 4px solid var(--green); border-radius: 8px; padding: 12px 16px; min-width: 150px; }}
.det-summary .box .lbl {{ font-size: 11px; color: var(--sub); text-transform: uppercase; letter-spacing: .5px; }}
.det-summary .box .val {{ font-size: 19px; font-weight: 700; margin-top: 4px; }}
.cup-row {{ cursor: pointer; }}
.cup-row td {{ font-weight: 600; }}
.cup-row:hover td {{ background: rgba(46,204,138,0.06); }}
.cup-toggle {{ display: inline-block; width: 14px; color: var(--green); font-weight: 700; }}
.prod-row td {{ background: #0d1f2d; font-weight: 400; font-size: 12px; color: var(--sub); padding-left: 28px; }}
.prod-head td {{ background: #0d1f2d; font-size: 10px; text-transform: uppercase; letter-spacing: .5px; color: var(--sub); font-weight: 700; }}
.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
.zero {{ color: #44607a; }}
th.sortable {{ cursor: pointer; user-select: none; white-space: nowrap; }}
th.sortable:hover {{ color: var(--green); }}
th.sortable .arr {{ color: var(--green); margin-left: 3px; }}
.pctcol {{ color: var(--mint); }}
.pctgreen {{ color: var(--green); font-weight: 600; }}
.pctorange {{ color: #FF9F1C; font-weight: 700; }}
.pctred {{ color: var(--red); font-weight: 700; }}
.just {{ color: var(--cyan); opacity: .85; }}
</style>
</head>
<body>

<div class="header">
  <div class="header-logo">A</div>
  <div class="header-text">
    <h1>Dashboard Auditoria de Descontos</h1>
    <p>Período: {periodo} &nbsp;|&nbsp; Gerado em: {gen_date}</p>
  </div>
</div>

<div class="tabs">
  <button class="tab-btn active" onclick="switchTab('manual')">⚠ Desconto Manual</button>
  <button class="tab-btn" onclick="switchTab('promo')">🏷 Desconto Promocional</button>
  <button class="tab-btn" onclick="switchTab('fidelidade')">⭐ Desconto Fidelidade</button>
  <button class="tab-btn" onclick="switchTab('detalhe')">🧾 Detalhamento (cupom a cupom)</button>
</div>

<div class="filters-bar" id="filtersBar"></div>

<div class="content">
  <div class="page active" id="page-manual">
    <div class="kpi-row" id="kpi-manual"></div>
    <div class="charts-grid" id="charts-manual-1"></div>
    <div class="charts-grid" id="charts-manual-2"></div>
  </div>

  <div class="page" id="page-promo">
    <div class="kpi-row" id="kpi-promo"></div>
    <div class="charts-grid" id="charts-promo-1"></div>
    <div class="charts-grid" id="charts-promo-2"></div>
    <div class="table-section">
      <div class="table-title">Top Campanhas Detalhadas</div>
      <div class="table-wrap" id="table-campanhas"></div>
    </div>
  </div>

  <div class="page" id="page-fidelidade">
    <div class="kpi-row" id="kpi-fidelidade"></div>
    <div class="charts-grid" id="charts-fid-1"></div>
    <div class="charts-grid" id="charts-fid-2"></div>
    <div class="table-section">
      <div class="table-title">Alertas F5 (Líquido Real ≤ 0)</div>
      <div class="table-wrap" id="table-f5"></div>
    </div>
  </div>

  <div class="page" id="page-detalhe">
    <div class="det-filters" id="det-filters">
      <div class="filter-group"><label>Loja (digite p/ buscar)</label><input type="text" id="det-loja" list="dl-det-loja" autocomplete="off" placeholder="— selecione/busque —"><datalist id="dl-det-loja"></datalist></div>
      <div class="filter-group"><label>Regional</label><select id="det-regional"><option value="">Todas</option></select></div>
      <div class="filter-group"><label>Praça</label><select id="det-praca"><option value="">Todas</option></select></div>
      <div class="filter-group"><label>Cluster</label><select id="det-cluster"><option value="">Todos</option></select></div>
      <div class="filter-group"><label>Mês (obrigatório)</label><select id="det-mes"><option value="">— selecione o mês —</option></select></div>
      <div class="filter-group"><label>Consultor (digite p/ buscar)</label><input type="text" id="det-cons" list="dl-det-cons" autocomplete="off" placeholder="Todos"><datalist id="dl-det-cons"></datalist></div>
      <div class="filter-group"><label>Nº Boleto / Cupom</label><input type="text" id="det-busca" placeholder="filtrar por boleto..."></div>
      <div class="filter-group"><label>Só com desconto</label><select id="det-comdesc"><option value="">Todos os cupons</option><option value="1">Apenas com desconto</option></select></div>
      <button class="btn-search" id="det-btn" onclick="buscarDetalhe()">Buscar cupons</button>
    </div>
    <div id="det-status" class="no-data">Selecione o <b>Mês</b> e uma <b>Loja</b> (ou Regional/Praça/Cluster) e clique em <b>Buscar cupons</b>.</div>
    <div id="det-summary" class="det-summary" style="display:none"></div>
    <div class="table-wrap" id="det-table" style="display:none"></div>
    <div class="pagination" id="det-pag"></div>
  </div>
</div>

<script src="painel_dados.js"></script>
<script>
// Dados principais (agregados) vêm de painel_dados.js -> window.PAINEL
const ROWS_HEADER = PAINEL.ROWS_HEADER;
const ROWS = PAINEL.ROWS;
const VIOLATIONS_DATA = PAINEL.VIOLATIONS;
const EXCEPTIONS_F = PAINEL.EXCEPTIONS_F;
const ITEMS = PAINEL.ITEMS;
const ITEMS_PROD = PAINEL.ITEMS_PROD;
const MES_LABELS = PAINEL.MES_LABELS;
const LK = PAINEL.LK;

const HDR = {{}};
ROWS_HEADER.forEach((h, i) => HDR[h] = i);

let activeFilters = {{mes:'', loja:'', consultor:'', canal:'', motivo2:'', campanha:'', dini:'', dfim:'', regional:'', praca:'', cluster:''}};
let currentPage = 'manual';
const chartInstances = {{}};
let violPage = 0;

function filteredRows() {{
  const f = activeFilters;
  return ROWS.filter(r => {{
    if (f.mes && r[HDR.mes] !== f.mes) return false;
    if (f.dini && r[HDR.data] < f.dini) return false;
    if (f.dfim && r[HDR.data] > f.dfim) return false;
    const lj = r[HDR.loja];
    if (f.loja !== '' && lj !== +f.loja) return false;
    if (f.regional && LK.loja_reg[lj] !== f.regional) return false;
    if (f.praca && LK.loja_praca[lj] !== f.praca) return false;
    if (f.cluster && LK.loja_cluster[lj] !== f.cluster) return false;
    if (f.consultor !== '' && r[HDR.consultor] !== +f.consultor) return false;
    if (f.canal !== '' && r[HDR.canal] !== +f.canal) return false;
    if (f.motivo2 !== '' && r[HDR.motivo2] !== +f.motivo2) return false;
    if (f.campanha !== '' && r[HDR.campanha] !== +f.campanha) return false;
    return true;
  }});
}}

function aggregate(rows, groupIdx, valueIdxs) {{
  const map = {{}};
  rows.forEach(r => {{
    const k = r[groupIdx];
    if (!map[k]) {{
      map[k] = {{key: k}};
      valueIdxs.forEach(v => map[k][v] = 0);
    }}
    valueIdxs.forEach(v => {{ map[k][v] += r[HDR[v]] || 0; }});
  }});
  return Object.values(map);
}}

// Ranking por % de desconto (desc/bruto), com piso de faturamento para evitar
// falsos líderes (ex.: 1 venda pequena com 100%). minBrutoFrac = fração do
// maior bruto do grupo usada como piso mínimo.
function topByPct(rows, groupIdx, descKey, topN, minBrutoFrac) {{
  const agg = aggregate(rows, groupIdx, [descKey, 'bruto']);
  const maxBruto = agg.reduce((m, x) => Math.max(m, x.bruto), 0);
  const floor = maxBruto * (minBrutoFrac || 0.01);
  return agg
    .filter(x => x.bruto >= floor && x.bruto > 0)
    .map(x => ({{key: x.key, pct: (x[descKey] / x.bruto) * 100, desc: x[descKey], bruto: x.bruto}}))
    .sort((a, b) => b.pct - a.pct)
    .slice(0, topN);
}}

// Ranking por % MANUAL = manual / (bruto - promo)  [Base_Manual do plano, Seção 1.5].
function topManualByPct(rows, groupIdx, topN, minBaseFrac) {{
  const agg = aggregate(rows, groupIdx, ['manual','bruto','promo']);
  const withBase = agg.map(x => ({{key: x.key, manual: x.manual, base: Math.max(0, x.bruto - x.promo)}}));
  const maxBase = withBase.reduce((m, x) => Math.max(m, x.base), 0);
  const floor = maxBase * (minBaseFrac || 0.02);
  return withBase
    .filter(x => x.base >= floor && x.base > 0 && x.manual > 0)
    .map(x => ({{key: x.key, pct: x.manual / x.base * 100, desc: x.manual, bruto: x.base}}))
    .sort((a, b) => b.pct - a.pct)
    .slice(0, topN);
}}

// Itens (rede) com maior % de desconto. metric = 'promo' ou 'manual'.
// Respeita o filtro de Mês; loja/consultor não se aplicam (visão de item da rede).
function topItensByPct(metric, topN) {{
  const mes = activeFilters.mes;
  const agg = {{}};  // prodIdx -> {{b, p, m}}
  for (let i = 0; i < ITEMS.length; i++) {{
    const it = ITEMS[i];
    if (mes && it[1] !== mes) continue;
    const k = it[0];
    if (!agg[k]) agg[k] = {{b:0, p:0, m:0}};
    agg[k].b += it[2]; agg[k].p += it[3]; agg[k].m += it[4];
  }}
  const arr = Object.keys(agg).map(k => {{
    const a = agg[k];
    // promo: base = bruto; manual: base = bruto - promo (Base_Manual do plano)
    const v = metric === 'promo' ? a.p : a.m;
    const base = metric === 'promo' ? a.b : Math.max(0, a.b - a.p);
    return {{name: ITEMS_PROD[k], pct: base>0 ? v/base*100 : 0, desc: v, bruto: base}};
  }});
  const maxB = arr.reduce((m,x) => Math.max(m, x.bruto), 0);
  return arr.filter(x => x.bruto >= maxB*0.02 && x.bruto > 0 && x.desc > 0)
            .sort((a,b) => b.pct - a.pct).slice(0, topN || 15);
}}

function brl(v) {{
  return (v||0).toLocaleString('pt-BR', {{style:'currency', currency:'BRL'}});
}}
function pct(a, b) {{
  if (!b) return '0,0%';
  return ((a/b)*100).toFixed(1).replace('.',',') + '%';
}}

function destroyChart(id) {{
  if (chartInstances[id]) {{
    chartInstances[id].destroy();
    delete chartInstances[id];
  }}
}}

const COLORS = ['#2ECC8A','#3355FF','#00AAFF','#4DE8AA','#FFB830','#FF4455','#a855f7','#f97316'];

// Descrição de cada regra de auditoria (mostrada na tabela e na legenda)
const REGRA_DESC = {{
  'A': 'Canal 11 (Parcerias/Convênio) com desconto MANUAL — manual proibido nesse canal',
  'C': 'Desconto manual SEM justificativa (Motivo 2 vazio ou genérico)',
  'F': 'Desconto total acima de 60% do valor bruto — teto absoluto da rede'
}};
function regraTexto(regra) {{
  return regra.split('+').map(r => REGRA_DESC[r] || r).join(' + ');
}}
// Converte filtro (índice) para o rótulo correspondente (p/ tabelas que guardam nomes)
function fName(idxStr, arr) {{ return idxStr === '' ? null : arr[+idxStr]; }}

// Remove consultores que são gestores (vendas no nome do gestor distorcem o % por consultor)
function semGestores(rows) {{
  const g = (LK.cons_gestor) || [];
  return rows.filter(r => !g[r[HDR.consultor]]);
}}

// Plugin Chart.js: desenha linha tracejada da MÉDIA no eixo de valor, com rótulo.
function avgLinePlugin(value, axis) {{
  return {{
    id: 'avgline',
    afterDraw(chart) {{
      if (value == null || !isFinite(value)) return;
      const {{ctx, chartArea, scales}} = chart;
      const sc = axis === 'y' ? scales.y : scales.x;
      if (!sc) return;
      ctx.save();
      ctx.strokeStyle = '#FF4455'; ctx.lineWidth = 2; ctx.setLineDash([5, 4]);
      const txt = 'Média ' + value.toFixed(1).replace('.', ',') + '%';
      ctx.font = 'bold 10px Arial'; ctx.fillStyle = '#FF4455';
      if (axis === 'y') {{
        const y = sc.getPixelForValue(value);
        ctx.beginPath(); ctx.moveTo(chartArea.left, y); ctx.lineTo(chartArea.right, y); ctx.stroke();
        ctx.setLineDash([]); ctx.fillText(txt, chartArea.left + 4, y - 4);
      }} else {{
        const x = sc.getPixelForValue(value);
        ctx.beginPath(); ctx.moveTo(x, chartArea.top); ctx.lineTo(x, chartArea.bottom); ctx.stroke();
        ctx.setLineDash([]); ctx.textAlign = 'left'; ctx.fillText(txt, Math.min(x + 4, chartArea.right - 70), chartArea.top + 11);
      }}
      ctx.restore();
    }}
  }};
}}

function makeChart(id, cfg) {{
  destroyChart(id);
  const canvas = document.getElementById(id);
  if (!canvas) return;
  chartInstances[id] = new Chart(canvas, cfg);
}}

function hbar(id, labels, data, color, isPct, extra, avg) {{
  // Se houver holder de rolagem, dimensiona a altura pelo nº de barras (≈22px cada)
  const holder = document.getElementById(id + '-holder');
  if (holder) holder.style.height = Math.max(340, labels.length * 22) + 'px';
  makeChart(id, {{
    type: 'bar',
    data: {{
      labels,
      datasets: [{{
        data,
        backgroundColor: color || '#2ECC8A',
        borderRadius: 4,
        borderSkipped: false,
      }}]
    }},
    plugins: (avg != null ? [avgLinePlugin(avg, 'x')] : []),
    options: {{
      indexAxis: 'y',
      maintainAspectRatio: false,
      responsive: true,
      plugins: {{
        legend: {{display: false}},
        tooltip: {{
          callbacks: {{
            label: ctx => {{
              if (isPct) {{
                const e = extra && extra[ctx.dataIndex];
                const base = e ? '  (' + brl(e.desc) + ' / ' + brl(e.bruto) + ')' : '';
                return ' ' + ctx.raw.toFixed(1).replace('.',',') + '%' + base;
              }}
              return ' ' + brl(ctx.raw);
            }}
          }}
        }}
      }},
      scales: {{
        x: {{
          ticks: {{color: '#8aa3b8', font:{{size:11}}, callback: v => isPct ? v.toFixed(0)+'%' : 'R$ ' + (v >= 1000000 ? (v/1000000).toFixed(1)+'M' : v >= 1000 ? (v/1000).toFixed(0)+'k' : v)}},
          grid: {{color: 'rgba(255,255,255,0.05)'}}
        }},
        y: {{
          ticks: {{color: '#fff', font:{{size:11}}}},
          grid: {{display: false}}
        }}
      }}
    }}
  }});
}}

function vbar(id, labels, data, color, yIsPct, avg) {{
  makeChart(id, {{
    type: 'bar',
    data: {{
      labels,
      datasets: [{{
        data,
        backgroundColor: color || '#00AAFF',
        borderRadius: 4,
        borderSkipped: false,
      }}]
    }},
    plugins: (avg != null ? [avgLinePlugin(avg, 'y')] : []),
    options: {{
      responsive: true,
      plugins: {{
        legend: {{display: false}},
        tooltip: {{
          callbacks: {{
            label: ctx => yIsPct ? ' ' + ctx.raw.toFixed(1) + '%' : ' ' + brl(ctx.raw)
          }}
        }}
      }},
      scales: {{
        x: {{
          ticks: {{color: '#8aa3b8', font:{{size:11}}, maxRotation: 35}},
          grid: {{display: false}}
        }},
        y: {{
          ticks: {{color: '#8aa3b8', font:{{size:11}}, callback: v => yIsPct ? v.toFixed(1)+'%' : (v >= 1000000 ? (v/1000000).toFixed(1)+'M' : v >= 1000 ? (v/1000).toFixed(0)+'k' : v)}},
          grid: {{color: 'rgba(255,255,255,0.05)'}}
        }}
      }}
    }}
  }});
}}

function lineChart(id, labels, data, color) {{
  makeChart(id, {{
    type: 'line',
    data: {{
      labels,
      datasets: [{{
        data,
        borderColor: color || '#2ECC8A',
        backgroundColor: (color || '#2ECC8A') + '22',
        fill: true,
        tension: 0.3,
        pointRadius: 5,
        pointBackgroundColor: color || '#2ECC8A',
      }}]
    }},
    options: {{
      responsive: true,
      plugins: {{
        legend: {{display: false}},
        tooltip: {{callbacks: {{label: ctx => ' ' + brl(ctx.raw)}}}}
      }},
      scales: {{
        x: {{ticks: {{color: '#8aa3b8', font:{{size:11}}}}, grid: {{display: false}}}},
        y: {{
          ticks: {{color: '#8aa3b8', font:{{size:11}}, callback: v => v >= 1000000 ? (v/1000000).toFixed(1)+'M' : v >= 1000 ? (v/1000).toFixed(0)+'k' : v}},
          grid: {{color: 'rgba(255,255,255,0.05)'}}
        }}
      }}
    }}
  }});
}}

function doughnut(id, labels, data) {{
  makeChart(id, {{
    type: 'doughnut',
    data: {{
      labels,
      datasets: [{{
        data,
        backgroundColor: COLORS,
        borderWidth: 2,
        borderColor: '#1C2B3A'
      }}]
    }},
    options: {{
      responsive: true,
      cutout: '60%',
      plugins: {{
        legend: {{
          position: 'right',
          labels: {{color: '#fff', font:{{size:11}}, boxWidth: 12, padding: 8}}
        }},
        tooltip: {{callbacks: {{label: ctx => ' ' + brl(ctx.raw)}}}}
      }}
    }}
  }});
}}

// Faixa de datas disponível (min/max) calculada uma vez
let DATA_MIN = '', DATA_MAX = '';
(function() {{
  for (let i = 0; i < ROWS.length; i++) {{
    const d = ROWS[i][HDR.data];
    if (!d) continue;
    if (!DATA_MIN || d < DATA_MIN) DATA_MIN = d;
    if (!DATA_MAX || d > DATA_MAX) DATA_MAX = d;
  }}
}})();

// Opções (índice -> rótulo) ordenadas por rótulo, para os campos pesquisáveis
function lkOptions(arr) {{
  return arr.map((label, i) => ({{i, label}}))
            .filter(o => o.label !== '' && o.label != null)
            .sort((a, b) => a.label.localeCompare(b.label));
}}

// Filter population
function populateFilters() {{
  const bar = document.getElementById('filtersBar');
  bar.innerHTML = '';

  // ── Mês (select simples) ──
  const mgrp = document.createElement('div'); mgrp.className = 'filter-group';
  mgrp.innerHTML = '<label>Mês</label>';
  const msel = document.createElement('select');
  msel.innerHTML = '<option value="">Todos</option>';
  Object.keys(MES_LABELS).sort().forEach(m => {{
    const op = document.createElement('option'); op.value = m; op.textContent = MES_LABELS[m];
    if (activeFilters.mes === m) op.selected = true;
    msel.appendChild(op);
  }});
  msel.onchange = () => {{ activeFilters.mes = msel.value; renderCurrentPage(); }};
  mgrp.appendChild(msel); bar.appendChild(mgrp);

  // ── Data início / fim (calendário) ──
  function dateInput(name, label) {{
    const g = document.createElement('div'); g.className = 'filter-group';
    g.innerHTML = `<label>${{label}}</label>`;
    const inp = document.createElement('input');
    inp.type = 'date'; inp.min = DATA_MIN; inp.max = DATA_MAX;
    if (activeFilters[name]) inp.value = activeFilters[name];
    inp.onchange = () => {{ activeFilters[name] = inp.value; renderCurrentPage(); }};
    g.appendChild(inp); bar.appendChild(g);
  }}
  dateInput('dini', 'Data início');
  dateInput('dfim', 'Data fim');

  // ── Regional / Praça / Cluster (selects simples) ──
  function selDim(name, label, opts) {{
    const g = document.createElement('div'); g.className = 'filter-group';
    g.innerHTML = `<label>${{label}}</label>`;
    const s = document.createElement('select');
    s.innerHTML = '<option value="">Todos</option>';
    opts.forEach(o => {{
      const op = document.createElement('option'); op.value = o; op.textContent = o;
      if (activeFilters[name] === o) op.selected = true;
      s.appendChild(op);
    }});
    s.onchange = () => {{ activeFilters[name] = s.value; renderCurrentPage(); }};
    g.appendChild(s); bar.appendChild(g);
  }}
  selDim('regional', 'Regional', LK.regionais || []);
  selDim('praca', 'Praça', LK.pracas || []);
  selDim('cluster', 'Cluster', LK.clusters || []);

  // ── Campos pesquisáveis (input + datalist), valor = índice ──
  let dlSeq = 0;
  function searchable(name, label, arr) {{
    const opts = lkOptions(arr);
    const g = document.createElement('div'); g.className = 'filter-group';
    g.innerHTML = `<label>${{label}}</label>`;
    const inp = document.createElement('input');
    inp.type = 'text'; inp.setAttribute('placeholder', 'Todos (digite p/ buscar)');
    inp.setAttribute('autocomplete', 'off');
    const dlId = 'dl-' + name + '-' + (dlSeq++);
    inp.setAttribute('list', dlId);
    const dl = document.createElement('datalist'); dl.id = dlId;
    opts.forEach(o => {{ const op = document.createElement('option'); op.value = o.label; dl.appendChild(op); }});
    // valor atual
    if (activeFilters[name] !== '') {{
      const cur = arr[+activeFilters[name]];
      if (cur != null) inp.value = cur;
    }}
    // mapa rótulo -> índice
    const byLabel = {{}}; opts.forEach(o => byLabel[o.label] = o.i);
    inp.onchange = () => {{
      const v = inp.value.trim();
      if (v === '') {{ activeFilters[name] = ''; }}
      else if (byLabel.hasOwnProperty(v)) {{ activeFilters[name] = String(byLabel[v]); }}
      else {{ activeFilters[name] = ''; inp.value = ''; }}  // texto não corresponde a opção
      renderCurrentPage();
    }};
    g.appendChild(inp); g.appendChild(dl); bar.appendChild(g);
  }}

  searchable('loja', 'Loja', LK.loja);
  searchable('consultor', 'Consultor', LK.cons);
  searchable('canal', 'Canal', LK.canal);
  searchable('motivo2', 'TAG', LK.motivo2);
  if (currentPage === 'promo') searchable('campanha', 'Campanha', LK.campanha);

  const clr = document.createElement('button');
  clr.className = 'btn-clear';
  clr.textContent = 'Limpar Filtros';
  clr.onclick = () => {{
    activeFilters = {{mes:'', loja:'', consultor:'', canal:'', motivo2:'', campanha:'', dini:'', dfim:'', regional:'', praca:'', cluster:''}};
    populateFilters();
    renderCurrentPage();
  }};
  clr.style.alignSelf = 'flex-end';
  bar.appendChild(clr);
}}

function cc(tag, cls, txt) {{
  const el = document.createElement(tag);
  if (cls) el.className = cls;
  if (txt !== undefined) el.textContent = txt;
  return el;
}}

function kpiCard(container, label, value, sub, colorClass) {{
  const card = cc('div', 'kpi-card ' + (colorClass||''));
  card.innerHTML = `<div class="kpi-label">${{label}}</div><div class="kpi-value">${{value}}</div><div class="kpi-sub">${{sub||''}}</div>`;
  container.appendChild(card);
}}

function chartArea(containerId, cards) {{
  const c = document.getElementById(containerId);
  c.innerHTML = '';
  cards.forEach(({{id, title, tall, scroll}}) => {{
    const card = cc('div', 'chart-card');
    if (scroll) {{
      card.innerHTML = `<div class="chart-title">${{title}}</div><div class="chart-scroll"><div class="chart-holder" id="${{id}}-holder"><canvas id="${{id}}"></canvas></div></div>`;
    }} else {{
      card.innerHTML = `<div class="chart-title">${{title}}</div><div class="chart-wrap${{tall?' tall':''}}"><canvas id="${{id}}"></canvas></div>`;
    }}
    c.appendChild(card);
  }});
}}

// ========== PAGE 1: MANUAL ==========
function renderManual() {{
  const rows = filteredRows();

  // KPIs
  const totalManual = rows.reduce((s, r) => s + r[HDR.manual], 0);
  const totalBruto = rows.reduce((s, r) => s + r[HDR.bruto], 0);
  const totalPromo = rows.reduce((s, r) => s + r[HDR.promo], 0);
  const baseManual = totalBruto - totalPromo;   // Base_Manual = Bruto - Promocional
  const cuponsManual = rows.filter(r => r[HDR.manual] > 0).reduce((s, r) => s + r[HDR.cnt], 0);
  const ticketManual = cuponsManual > 0 ? totalManual / cuponsManual : 0;

  const kpiEl = document.getElementById('kpi-manual');
  kpiEl.innerHTML = '';
  kpiCard(kpiEl, 'Total R$ Manual', brl(totalManual), 'Soma descontos manuais', '');
  kpiCard(kpiEl, '% Manual (base pós-promo)', pct(totalManual, baseManual), 'Manual / (Bruto − Promocional)', 'yellow');
  kpiCard(kpiEl, 'Itens c/ Manual', cuponsManual.toLocaleString('pt-BR'), 'Itens com desconto manual', 'blue');
  kpiCard(kpiEl, 'Manual Médio / Item', brl(ticketManual), 'R$ manual ÷ itens c/ manual', 'green');

  // Charts row 1
  chartArea('charts-manual-1', [
    {{id:'ch-m-consult', title:'Consultores por % Manual (Manual ÷ Bruto−Promo) — role p/ ver todos', scroll:true}},
    {{id:'ch-m-lojas', title:'Lojas por % Manual (Manual ÷ Bruto−Promo) — role p/ ver todas', scroll:true}}
  ]);

  const avgManual = baseManual > 0 ? totalManual / baseManual * 100 : 0;

  // Consultores por % manual (base pós-promo) — SEM gestores, todos, do menor p/ maior
  const byConsult = topManualByPct(semGestores(rows), HDR.consultor, 600, 0);
  hbar('ch-m-consult', byConsult.map(x=>LK.cons[x.key]), byConsult.map(x=>x.pct), '#2ECC8A', true, byConsult, avgManual);

  // Lojas por % manual (base pós-promo) — todas, do menor p/ maior
  const byLoja = topManualByPct(rows, HDR.loja, 600, 0);
  hbar('ch-m-lojas', byLoja.map(x=>LK.loja[x.key]), byLoja.map(x=>x.pct), '#3355FF', true, byLoja, avgManual);

  // Charts row 2
  chartArea('charts-manual-2', [
    {{id:'ch-m-itens', title:'Itens por % Manual (rede) — role p/ ver todos', scroll:true}},
    {{id:'ch-m-motivo', title:'Distribuição por TAG (Top 8)'}},
    {{id:'ch-m-canal', title:'% Manual por Canal (Manual ÷ Bruto−Promo)'}},
    {{id:'ch-m-mes', title:'Evolução por Mês'}}
  ]);

  // Itens por % manual (rede, base pós-promo) — todos, do menor p/ maior
  const itM = topItensByPct('manual', 600);
  hbar('ch-m-itens', itM.map(x=>x.name), itM.map(x=>x.pct), '#2ECC8A', true, itM, avgManual);

  // Motivo2
  const byMotivo = aggregate(rows.filter(r => r[HDR.manual] > 0), HDR.motivo2, ['manual'])
    .sort((a,b) => b.manual - a.manual).slice(0, 8);
  doughnut('ch-m-motivo', byMotivo.map(x => LK.motivo2[x.key] || '(sem motivo)'), byMotivo.map(x => x.manual));

  // Canal % (base pós-promo)
  const byCanal = aggregate(rows.filter(r => r[HDR.bruto] > 0), HDR.canal, ['manual','bruto','promo'])
    .sort((a,b) => b.bruto - a.bruto).slice(0, 10);
  const canalPct = byCanal.map(x => {{ const base = x.bruto - x.promo; return base > 0 ? (x.manual/base)*100 : 0; }});
  vbar('ch-m-canal', byCanal.map(x => LK.canal[x.key]), canalPct, '#00AAFF', true, avgManual);

  // Mes line
  const byMes = aggregate(rows, HDR.mes, ['manual']).sort((a,b) => a.key < b.key ? -1 : 1);
  lineChart('ch-m-mes', byMes.map(x => MES_LABELS[x.key]||x.key), byMes.map(x => x.manual), '#2ECC8A');

}}

let f60Page = 0;
function renderF60Table(page) {{
  f60Page = page;
  const f = activeFilters;
  const fLoja = fName(f.loja, LK.loja), fCons = fName(f.consultor, LK.cons), fCanal = fName(f.canal, LK.canal);
  let data = EXCEPTIONS_F.filter(v => {{
    if (fLoja && v.loja !== fLoja) return false;
    if (fCons && v.consultor !== fCons) return false;
    if (fCanal && v.canal !== fCanal) return false;
    if (f.regional && v.regional !== f.regional) return false;
    if (f.praca && v.praca !== f.praca) return false;
    if (f.cluster && v.cluster !== f.cluster) return false;
    if (f.mes && (v.data||'').slice(0,7) !== f.mes) return false;
    if (f.dini && v.data < f.dini) return false;
    if (f.dfim && v.data > f.dfim) return false;
    return true;
  }});

  const PER_PAGE = 20;
  const totalPages = Math.max(1, Math.ceil(data.length / PER_PAGE));
  const slice = data.slice(page * PER_PAGE, (page+1) * PER_PAGE);

  const wrap = document.getElementById('table-f60');
  const legenda = `<div class="legenda"><b>Regra F:</b> ${{REGRA_DESC['F']}}</div>`;
  if (!slice.length) {{
    wrap.innerHTML = legenda + '<div class="no-data">Nenhuma exceção de Regra F (desconto &gt; 60%) com os filtros selecionados.</div>';
    document.getElementById('pag-f60').innerHTML = '';
    return;
  }}

  let html = legenda + '<table><thead><tr><th>Regra</th><th>Loja</th><th>Consultor</th><th>Canal</th><th>Origem Predom.</th><th class="num">R$ Bruto</th><th class="num">R$ Desconto</th><th class="num">% Desconto</th></tr></thead><tbody>';
  slice.forEach(v => {{
    html += `<tr><td><span class="tag tag-red" title="${{REGRA_DESC['F']}}">F · &gt;60%</span></td><td>${{v.loja}}</td><td>${{v.consultor}}</td><td>${{v.canal}}</td><td>${{v.origem}}</td><td class="num">${{brl(v.bruto)}}</td><td class="num" style="color:var(--red)">${{brl(v.desc_total)}}</td><td class="num"><span class="tag tag-red">${{v.pct.toFixed(1).replace('.',',')}}%</span></td></tr>`;
  }});
  html += '</tbody></table>';
  wrap.innerHTML = html;

  const pag = document.getElementById('pag-f60');
  pag.innerHTML = `<span>${{data.length}} exceções</span>`;
  for (let p = 0; p < totalPages; p++) {{
    const btn = document.createElement('button');
    btn.textContent = p + 1;
    if (p === page) btn.classList.add('active');
    btn.onclick = () => renderF60Table(p);
    pag.appendChild(btn);
  }}
}}

function renderViolationsTable(page) {{
  violPage = page;
  const f = activeFilters;
  const fLoja = fName(f.loja, LK.loja), fCons = fName(f.consultor, LK.cons),
        fCanal = fName(f.canal, LK.canal), fMot = fName(f.motivo2, LK.motivo2);
  let data = VIOLATIONS_DATA.filter(v => {{
    if (fLoja && v.loja !== fLoja) return false;
    if (fCons && v.consultor !== fCons) return false;
    if (fCanal && v.canal !== fCanal) return false;
    if (fMot && v.motivo2 !== fMot) return false;
    if (f.regional && v.regional !== f.regional) return false;
    if (f.praca && v.praca !== f.praca) return false;
    if (f.cluster && v.cluster !== f.cluster) return false;
    if (f.mes && (v.data||'').slice(0,7) !== f.mes) return false;
    if (f.dini && v.data < f.dini) return false;
    if (f.dfim && v.data > f.dfim) return false;
    return true;
  }});

  const PER_PAGE = 20;
  const totalPages = Math.max(1, Math.ceil(data.length / PER_PAGE));
  const slice = data.slice(page * PER_PAGE, (page+1) * PER_PAGE);

  const wrap = document.getElementById('table-violations');
  const legenda = `<div class="legenda"><b>Regra A:</b> ${{REGRA_DESC['A']}} &nbsp;·&nbsp; <b>Regra C:</b> ${{REGRA_DESC['C']}}</div>`;
  if (!slice.length) {{
    wrap.innerHTML = legenda + '<div class="no-data">Nenhuma violação encontrada com os filtros selecionados.</div>';
    document.getElementById('pag-violations').innerHTML = '';
    return;
  }}

  let html = legenda + '<table><thead><tr><th>Regra</th><th>O que se refere</th><th>Loja</th><th>Consultor</th><th>Canal</th><th class="num">R$ Manual</th><th class="num">R$ Bruto</th><th>Motivo 2</th></tr></thead><tbody>';
  slice.forEach(v => {{
    const tagCls = v.regra.includes('A') ? 'tag-red' : 'tag-yellow';
    const lbl = v.regra.split('+').map(r => 'Regra ' + r).join(' + ');
    html += `<tr><td><span class="tag ${{tagCls}}" title="${{regraTexto(v.regra)}}">${{lbl}}</span></td><td><small>${{regraTexto(v.regra)}}</small></td><td>${{v.loja}}</td><td>${{v.consultor}}</td><td>${{v.canal}}</td><td class="num">${{brl(v.manual)}}</td><td class="num">${{brl(v.bruto)}}</td><td>${{v.motivo2||'—'}}</td></tr>`;
  }});
  html += '</tbody></table>';
  wrap.innerHTML = html;

  const pag = document.getElementById('pag-violations');
  pag.innerHTML = `<span>${{data.length}} violações</span>`;
  for (let p = 0; p < totalPages; p++) {{
    const btn = document.createElement('button');
    btn.textContent = p + 1;
    if (p === page) btn.classList.add('active');
    btn.onclick = () => renderViolationsTable(p);
    pag.appendChild(btn);
  }}
}}

// ========== PAGE 2: PROMO ==========
function renderPromo() {{
  const rows = filteredRows();

  const totalPromo = rows.reduce((s, r) => s + r[HDR.promo], 0);
  const totalBruto = rows.reduce((s, r) => s + r[HDR.bruto], 0);
  const cuponsPromo = rows.filter(r => r[HDR.promo] > 0).reduce((s, r) => s + r[HDR.cnt], 0);
  const qtdCamp = new Set(rows.filter(r => r[HDR.promo] > 0 && LK.campanha[r[HDR.campanha]]).map(r => r[HDR.campanha])).size;

  const kpiEl = document.getElementById('kpi-promo');
  kpiEl.innerHTML = '';
  kpiCard(kpiEl, 'Total R$ Promo', brl(totalPromo), 'Soma descontos promocionais', '');
  kpiCard(kpiEl, '% do Bruto', pct(totalPromo, totalBruto), 'Promo / Valor Bruto', 'yellow');
  kpiCard(kpiEl, 'Qtd Cupons c/ Promo', cuponsPromo.toLocaleString('pt-BR'), 'Registros com desc. promo', 'blue');
  kpiCard(kpiEl, 'Qtd Campanhas', qtdCamp.toLocaleString('pt-BR'), 'Campanhas distintas', 'cyan');

  chartArea('charts-promo-1', [
    {{id:'ch-p-camp', title:'Campanhas por % Promo (sobre bruto) — role p/ ver todas', scroll:true}},
    {{id:'ch-p-lojas', title:'Lojas por % Promo (sobre bruto) — role p/ ver todas', scroll:true}}
  ]);

  const avgPromo = totalBruto > 0 ? totalPromo / totalBruto * 100 : 0;

  const byCamp = topByPct(rows.filter(r => LK.campanha[r[HDR.campanha]]), HDR.campanha, 'promo', 600, 0);
  hbar('ch-p-camp', byCamp.map(x => LK.campanha[x.key]), byCamp.map(x => x.pct), '#2ECC8A', true, byCamp, avgPromo);

  const byLoja = topByPct(rows, HDR.loja, 'promo', 600, 0);
  hbar('ch-p-lojas', byLoja.map(x => LK.loja[x.key]), byLoja.map(x => x.pct), '#3355FF', true, byLoja, avgPromo);

  chartArea('charts-promo-2', [
    {{id:'ch-p-itens', title:'Itens por % Promocional (rede) — role p/ ver todos', scroll:true}},
    {{id:'ch-p-mes', title:'Evolução por Mês'}},
    {{id:'ch-p-canal', title:'Promo por Canal'}}
  ]);

  // Itens por % promocional (rede) — todos, do menor p/ maior
  const itP = topItensByPct('promo', 600);
  hbar('ch-p-itens', itP.map(x=>x.name), itP.map(x=>x.pct), '#2ECC8A', true, itP, avgPromo);

  const byMes = aggregate(rows, HDR.mes, ['promo']).sort((a,b) => a.key < b.key ? -1 : 1);
  lineChart('ch-p-mes', byMes.map(x => MES_LABELS[x.key]||x.key), byMes.map(x => x.promo), '#00AAFF');

  const byCanal = aggregate(rows.filter(r => r[HDR.promo] > 0), HDR.canal, ['promo','bruto'])
    .sort((a,b) => b.promo - a.promo).slice(0, 10);
  vbar('ch-p-canal', byCanal.map(x => LK.canal[x.key]), byCanal.map(x => x.promo), '#4DE8AA');

  // Campanhas table
  const campData = aggregate(rows.filter(r => r[HDR.promo] > 0 && LK.campanha[r[HDR.campanha]]), HDR.campanha, ['promo','bruto','cnt'])
    .sort((a,b) => b.promo - a.promo).slice(0, 50);

  let html = '<table><thead><tr><th>Campanha</th><th class="num">R$ Promo</th><th class="num">Qtd</th><th class="num">% Bruto</th></tr></thead><tbody>';
  campData.forEach(v => {{
    html += `<tr><td>${{LK.campanha[v.key]}}</td><td class="num">${{brl(v.promo)}}</td><td class="num">${{v.cnt.toLocaleString('pt-BR')}}</td><td class="num">${{pct(v.promo,v.bruto)}}</td></tr>`;
  }});
  html += '</tbody></table>';
  document.getElementById('table-campanhas').innerHTML = html || '<div class="no-data">Nenhum dado.</div>';
}}

// ========== PAGE 3: FIDELIDADE ==========
function renderFidelidade() {{
  const rows = filteredRows();

  const totalFid = rows.reduce((s, r) => s + r[HDR.fidelidade], 0);
  const totalBruto = rows.reduce((s, r) => s + r[HDR.bruto], 0);
  const cuponsFid = rows.filter(r => r[HDR.fidelidade] > 0).reduce((s, r) => s + r[HDR.cnt], 0);
  const totalF5 = rows.reduce((s, r) => s + r[HDR.f5_cnt], 0);

  const kpiEl = document.getElementById('kpi-fidelidade');
  kpiEl.innerHTML = '';
  kpiCard(kpiEl, 'Total R$ Fidelidade', brl(totalFid), 'Soma descontos fidelidade', '');
  kpiCard(kpiEl, '% do Bruto', pct(totalFid, totalBruto), 'Fidelidade / Valor Bruto', 'yellow');
  kpiCard(kpiEl, 'Qtd Cupons c/ Fidelidade', cuponsFid.toLocaleString('pt-BR'), 'Registros com desc. fidelidade', 'blue');
  kpiCard(kpiEl, 'Alertas F5', totalF5.toLocaleString('pt-BR'), 'Líquido Real ≤ 0', 'red');

  chartArea('charts-fid-1', [
    {{id:'ch-f-lojas', title:'Lojas por % Fidelidade (sobre bruto) — role p/ ver todas', scroll:true}},
    {{id:'ch-f-consult', title:'Consultores por % Fidelidade (sobre bruto) — role p/ ver todos', scroll:true}}
  ]);

  const avgFid = totalBruto > 0 ? totalFid / totalBruto * 100 : 0;

  const byLoja = topByPct(rows, HDR.loja, 'fidelidade', 600, 0);
  hbar('ch-f-lojas', byLoja.map(x => LK.loja[x.key]), byLoja.map(x => x.pct), '#2ECC8A', true, byLoja, avgFid);

  // Consultores SEM gestores — todos, do menor p/ maior
  const byConsult = topByPct(semGestores(rows), HDR.consultor, 'fidelidade', 600, 0);
  hbar('ch-f-consult', byConsult.map(x => LK.cons[x.key]), byConsult.map(x => x.pct), '#3355FF', true, byConsult, avgFid);

  chartArea('charts-fid-2', [
    {{id:'ch-f-mes', title:'Evolução por Mês'}},
    {{id:'ch-f-canal', title:'% Fidelidade por Canal'}}
  ]);

  const byMes = aggregate(rows, HDR.mes, ['fidelidade']).sort((a,b) => a.key < b.key ? -1 : 1);
  lineChart('ch-f-mes', byMes.map(x => MES_LABELS[x.key]||x.key), byMes.map(x => x.fidelidade), '#00AAFF');

  const byCanal = aggregate(rows.filter(r => r[HDR.bruto] > 0), HDR.canal, ['fidelidade','bruto'])
    .sort((a,b) => b.bruto - a.bruto).slice(0, 10);
  const canalPct = byCanal.map(x => x.bruto > 0 ? (x.fidelidade/x.bruto)*100 : 0);
  vbar('ch-f-canal', byCanal.map(x => LK.canal[x.key]), canalPct, '#4DE8AA', true, avgFid);

  // F5 table
  const f5Data = aggregate(rows.filter(r => r[HDR.f5_cnt] > 0), HDR.loja, ['f5_cnt','fidelidade','bruto'])
    .sort((a,b) => b.f5_cnt - a.f5_cnt).slice(0, 50);

  const legF5 = '<div class="legenda"><b>Regra F5:</b> Líquido real do item ≤ 0 após Promo + Manual + Fidelidade — margem evaporada.</div>';
  let html = legF5 + '<table><thead><tr><th>Loja</th><th class="num">Alertas F5</th><th class="num">R$ Fidelidade</th><th class="num">R$ Bruto</th></tr></thead><tbody>';
  f5Data.forEach(v => {{
    html += `<tr><td>${{LK.loja[v.key]}}</td><td class="num">${{v.f5_cnt.toLocaleString('pt-BR')}}</td><td class="num">${{brl(v.fidelidade)}}</td><td class="num">${{brl(v.bruto)}}</td></tr>`;
  }});
  html += '</tbody></table>';
  document.getElementById('table-f5').innerHTML = html || '<div class="no-data">Nenhum alerta F5.</div>';
}}

// ========== ABA: DETALHAMENTO (lazy-load) ==========
let DET_LOADED = false, DET_ROWS = [], detPage = 0;

function loadScriptOnce(src, cb, onerr) {{
  const s = document.createElement('script');
  s.src = src; s.onload = cb;
  s.onerror = onerr || (() => {{}});
  document.body.appendChild(s);
}}

function loadDetalhe(cb) {{
  if (DET_LOADED) {{ cb && cb(); return; }}
  const st = document.getElementById('det-status');
  st.style.display = 'block';
  st.innerHTML = 'Carregando cadastro de lojas/produtos...';
  loadScriptOnce('detalhe_base.js', () => {{
    DET_LOADED = true;
    initDetFilters();
    st.innerHTML = 'Selecione o <b>Mês</b> e uma <b>Loja</b> (ou Regional/Praça/Cluster) e clique em <b>Buscar cupons</b>.';
    cb && cb();
  }}, () => {{
    st.innerHTML = '⚠ Não foi possível carregar <b>detalhe_base.js</b>. Mantenha os arquivos na mesma pasta.';
  }});
}}

// Carrega o arquivo de cupons de um mês sob demanda
function loadMonth(mes, cb) {{
  if (window.DET_M && window.DET_M[mes]) {{ cb(); return; }}
  const st = document.getElementById('det-status');
  st.style.display = 'block';
  st.innerHTML = 'Carregando cupons de ' + (window.DET_BASE.mes_labels[mes] || mes) + '...';
  loadScriptOnce('detalhe_' + mes + '.js', () => cb(), () => {{
    st.innerHTML = '⚠ Não foi possível carregar os cupons de ' + mes + '.';
  }});
}}

let detLojaByName = {{}}, detConsByName = {{}};
function initDetFilters() {{
  const D = window.DET_BASE;
  // Loja (datalist pesquisável) — value = nome, mapeia p/ índice
  const dlLoja = document.getElementById('dl-det-loja');
  detLojaByName = {{}};
  D.lojas.map((n,i)=>({{n,i}})).sort((a,b)=>a.n.localeCompare(b.n)).forEach(o => {{
    const op = document.createElement('option'); op.value = o.n; dlLoja.appendChild(op);
    detLojaByName[o.n] = o.i;
  }});
  // Consultor (datalist pesquisável)
  const dlCons = document.getElementById('dl-det-cons');
  detConsByName = {{}};
  D.cons.map((n,i)=>({{n,i}})).sort((a,b)=>a.n.localeCompare(b.n)).forEach(o => {{
    const op = document.createElement('option'); op.value = o.n; dlCons.appendChild(op);
    detConsByName[o.n] = o.i;
  }});
  // Mês (obrigatório) — usa os meses disponíveis; pré-seleciona o mais recente
  const mesSel = document.getElementById('det-mes');
  const meses = (D.meses || Object.keys(D.mes_labels)).slice().sort();
  meses.forEach(m => {{
    const op = document.createElement('option'); op.value = m; op.textContent = D.mes_labels[m] || m; mesSel.appendChild(op);
  }});
  if (meses.length) mesSel.value = meses[meses.length - 1];
  // Regional / Praça / Cluster
  function fill(id, arr) {{
    const s = document.getElementById(id);
    (arr||[]).forEach(v => {{ const op = document.createElement('option'); op.value = v; op.textContent = v; s.appendChild(op); }});
  }}
  fill('det-regional', D.regionais);
  fill('det-praca', D.pracas);
  fill('det-cluster', D.clusters);
}}

function buscarDetalhe() {{
  const D = window.DET_BASE;
  if (!D) {{ loadDetalhe(buscarDetalhe); return; }}
  const mesV = document.getElementById('det-mes').value;
  if (!mesV) {{ alert('Selecione o Mês primeiro.'); return; }}
  const lojaTxt = document.getElementById('det-loja').value.trim();
  const regional = document.getElementById('det-regional').value;
  const praca = document.getElementById('det-praca').value;
  const cluster = document.getElementById('det-cluster').value;
  let li = -1;
  if (lojaTxt !== '') {{
    if (!detLojaByName.hasOwnProperty(lojaTxt)) {{ alert('Loja não encontrada. Escolha uma opção da lista.'); return; }}
    li = detLojaByName[lojaTxt];
  }}
  if (li < 0 && !regional && !praca && !cluster) {{
    alert('Selecione uma Loja (ou Regional/Praça/Cluster) para filtrar.');
    return;
  }}
  const consTxt = document.getElementById('det-cons').value.trim();
  const ci = (consTxt !== '' && detConsByName.hasOwnProperty(consTxt)) ? detConsByName[consTxt] : -1;
  const busca = document.getElementById('det-busca').value.trim();
  const comdesc = document.getElementById('det-comdesc').value;

  loadMonth(mesV, () => {{
    const cup = (window.DET_M && window.DET_M[mesV]) || [];
    DET_ROWS = cup.filter(c => {{
      const lj = c[1];
      if (li >= 0 && lj !== li) return false;
      if (regional && D.loja_reg[lj] !== regional) return false;
      if (praca && D.loja_praca[lj] !== praca) return false;
      if (cluster && D.loja_cluster[lj] !== cluster) return false;
      if (ci >= 0 && c[3] !== ci) return false;
      if (busca && String(c[0]).indexOf(busca) < 0) return false;
      if (comdesc === '1' && (c[6]+c[7]+c[8]) <= 0) return false;
      return true;
    }});
    detPage = 0;
    applyDetSort();
    renderDetSummary();
    if (DET_ROWS.length) renderDetTable(0);
  }});
}}

function detNum(v) {{ return v>0 ? brl(v) : '<span class="zero">—</span>'; }}

function renderDetSummary() {{
  const sm = document.getElementById('det-summary');
  const st = document.getElementById('det-status');
  let b=0,p=0,m=0,f=0;
  DET_ROWS.forEach(c => {{ b+=c[5]; p+=c[6]; m+=c[7]; f+=c[8]; }});
  if (!DET_ROWS.length) {{
    sm.style.display='none';
    document.getElementById('det-table').style.display='none';
    document.getElementById('det-pag').innerHTML='';
    st.style.display='block';
    st.textContent='Nenhum cupom encontrado para os filtros selecionados.';
    return;
  }}
  st.style.display='none';
  sm.style.display='flex';
  sm.innerHTML = `
    <div class="box"><div class="lbl">Cupons</div><div class="val">${{DET_ROWS.length.toLocaleString('pt-BR')}}</div></div>
    <div class="box" style="border-left-color:var(--text)"><div class="lbl">Bruto</div><div class="val">${{brl(b)}}</div></div>
    <div class="box" style="border-left-color:var(--cyan)"><div class="lbl">Promocional</div><div class="val">${{brl(p)}}</div></div>
    <div class="box" style="border-left-color:var(--blue)"><div class="lbl">Manual</div><div class="val">${{brl(m)}}</div></div>
    <div class="box" style="border-left-color:var(--mint)"><div class="lbl">Fidelidade</div><div class="val">${{brl(f)}}</div></div>`;
}}

// Colunas ordenáveis do nível cupom (n = id lógico para sort)
const DET_COLS = [
  {{l:'Boleto', n:0}}, {{l:'Loja', n:12}}, {{l:'Data', n:1}}, {{l:'Consultor', n:2}}, {{l:'Itens', n:3, num:1}},
  {{l:'Bruto', n:4, num:1}}, {{l:'Promo', n:5, num:1}}, {{l:'Manual', n:6, num:1}}, {{l:'Fidelid.', n:7, num:1}},
  {{l:'%Promo', n:8, num:1, pct:1}}, {{l:'%Manual', n:9, num:1, pct:1}}, {{l:'%Total', n:10, num:1, pct:1}},
  {{l:'Líquido', n:11, num:1}}
];
let detSort = {{col: 10, dir: -1}};  // padrão: %Total desc

function cupSortVal(c, col) {{
  const D = window.DET_BASE;
  const b = c[5];
  switch(col) {{
    case 0: return parseInt(c[0], 10) || 0;
    case 12: return D.lojas[c[1]] || '';
    case 1: {{ const p = (D.datas[c[2]]||'').split('/'); return p.length===3 ? (+p[2]*10000 + +p[1]*100 + +p[0]) : 0; }}
    case 2: return D.cons[c[3]] || '';
    case 3: return c[9].length;
    case 4: return b;
    case 5: return c[6];
    case 6: return c[7];
    case 7: return c[8];
    case 8: return b>0 ? c[6]/b : 0;
    case 9: {{ const bm = b - c[6]; return bm>0 ? c[7]/bm : 0; }}
    case 10: return b>0 ? (c[6]+c[7]+c[8])/b : 0;
    case 11: return c[5]-c[6]-c[7]-c[8];
  }}
  return 0;
}}

function applyDetSort() {{
  const col = detSort.col, dir = detSort.dir;
  DET_ROWS.sort((a,b) => {{
    const va = cupSortVal(a, col), vb = cupSortVal(b, col);
    if (typeof va === 'string') return va.localeCompare(vb) * dir;
    return (va - vb) * dir;
  }});
}}

function sortDet(col) {{
  if (detSort.col === col) detSort.dir *= -1;
  else {{ detSort.col = col; detSort.dir = (col <= 2 ? 1 : -1); }}
  applyDetSort();
  renderDetTable(0);
}}

function detPct(v) {{
  if (v <= 0) return '<span class="zero">—</span>';
  // 0–30% verde | >30%–<50% laranja | >50% vermelho
  const cls = v > 0.50 ? 'pctred' : (v > 0.30 ? 'pctorange' : 'pctgreen');
  return '<span class="' + cls + '">' + (v*100).toFixed(1).replace('.', ',') + '%</span>';
}}

function renderDetTable(page) {{
  detPage = page;
  const D = window.DET_BASE;
  const PER = 25;
  const total = DET_ROWS.length;
  const pages = Math.max(1, Math.ceil(total/PER));
  const slice = DET_ROWS.slice(page*PER, (page+1)*PER);
  const tbl = document.getElementById('det-table');
  tbl.style.display = 'block';
  let head = '<th></th>';
  DET_COLS.forEach(col => {{
    const arr = detSort.col === col.n ? `<span class="arr">${{detSort.dir<0?'▼':'▲'}}</span>` : '';
    head += `<th class="sortable${{col.num?' num':''}}" onclick="sortDet(${{col.n}})">${{col.l}}${{arr}}</th>`;
  }});
  let h = '<div class="legenda" style="margin-bottom:8px">%Promo e %Total sobre o <b>Valor Bruto</b>. <b>%Manual = Manual ÷ (Bruto − Promocional)</b> — base pós-promocional, conforme regra de alçada.</div>';
  h += '<table><thead><tr>' + head + '</tr></thead><tbody>';
  slice.forEach((c,k) => {{
    const b = c[5], liq = b-c[6]-c[7]-c[8];
    const baseM = b - c[6];  // Base_Manual = bruto - promo
    const pPromo = b>0 ? c[6]/b : 0, pMan = baseM>0 ? c[7]/baseM : 0, pTot = b>0 ? (c[6]+c[7]+c[8])/b : 0;
    h += `<tr class="cup-row" onclick="toggleCup(${{k}})"><td><span class="cup-toggle" id="tg-${{k}}">▶</span></td><td>${{c[0]}}</td><td>${{D.lojas[c[1]]}}</td><td>${{D.datas[c[2]]}}</td><td>${{D.cons[c[3]]}}</td><td class="num">${{c[9].length}}</td><td class="num">${{detNum(b)}}</td><td class="num">${{detNum(c[6])}}</td><td class="num">${{detNum(c[7])}}</td><td class="num">${{detNum(c[8])}}</td><td class="num">${{detPct(pPromo)}}</td><td class="num">${{detPct(pMan)}}</td><td class="num">${{detPct(pTot)}}</td><td class="num">${{detNum(liq)}}</td></tr>`;
    h += `<tr class="prod-head pc-${{k}}" style="display:none"><td></td><td colspan="4">Produto (SKU)</td><td class="num">Qtd</td><td class="num">Bruto</td><td class="num">Promo</td><td class="num">Manual</td><td class="num">Fidelid.</td><td class="num">%Promo</td><td class="num">%Manual</td><td class="num">%Total</td><td class="num">Líquido</td></tr>`;
    c[9].forEach(it => {{
      const ib = it[3], pliq = ib-it[4]-it[5]-it[6];
      const ibaseM = ib - it[4];  // Base_Manual = bruto - promo
      const ipPromo = ib>0 ? it[4]/ib : 0, ipMan = ibaseM>0 ? it[5]/ibaseM : 0, ipTot = ib>0 ? (it[4]+it[5]+it[6])/ib : 0;
      const dupBadge = (it[7] > 1) ? ` <span class="tag tag-red" title="Item escaneado ${{it[7]}}x em vez de usar a quantidade">⟳ ${{it[7]}}x</span>` : '';
      const jt = [];
      if (it[5] > 0 && D.mot && D.mot[it[8]]) jt.push('<b>TAG:</b> ' + D.mot[it[8]]);
      if (it[4] > 0 && D.camp && D.camp[it[9]]) jt.push('<b>Campanha:</b> ' + D.camp[it[9]]);
      const justHtml = jt.length ? `<br><small class="just">${{jt.join(' &nbsp;·&nbsp; ')}}</small>` : '';
      h += `<tr class="prod-row pc-${{k}}" style="display:none"><td></td><td colspan="4">${{D.prod[it[1]]}} <small style="opacity:.55">(${{it[0]}})</small>${{dupBadge}}${{justHtml}}</td><td class="num">${{it[2]}}</td><td class="num">${{detNum(ib)}}</td><td class="num">${{detNum(it[4])}}</td><td class="num">${{detNum(it[5])}}</td><td class="num">${{detNum(it[6])}}</td><td class="num">${{detPct(ipPromo)}}</td><td class="num">${{detPct(ipMan)}}</td><td class="num">${{detPct(ipTot)}}</td><td class="num">${{detNum(pliq)}}</td></tr>`;
    }});
  }});
  h += '</tbody></table>';
  tbl.innerHTML = h;

  const pag = document.getElementById('det-pag');
  pag.innerHTML = '';
  const info = document.createElement('span');
  info.textContent = `${{total.toLocaleString('pt-BR')}} cupons · página ${{page+1}} de ${{pages}}`;
  const prev = document.createElement('button'); prev.textContent = '‹ Anterior';
  prev.onclick = () => {{ if (detPage>0) renderDetTable(detPage-1); }};
  const next = document.createElement('button'); next.textContent = 'Próxima ›';
  next.onclick = () => {{ if (detPage<pages-1) renderDetTable(detPage+1); }};
  pag.appendChild(info); pag.appendChild(prev); pag.appendChild(next);
}}

function toggleCup(k) {{
  const rows = document.querySelectorAll('.pc-'+k);
  const tg = document.getElementById('tg-'+k);
  const show = rows.length && rows[0].style.display === 'none';
  rows.forEach(r => r.style.display = show ? 'table-row' : 'none');
  if (tg) tg.textContent = show ? '▼' : '▶';
}}

function renderCurrentPage() {{
  if (currentPage === 'manual') renderManual();
  else if (currentPage === 'promo') renderPromo();
  else if (currentPage === 'fidelidade') renderFidelidade();
}}

function switchTab(page) {{
  currentPage = page;
  const pages = ['manual','promo','fidelidade','detalhe'];
  document.querySelectorAll('.tab-btn').forEach((b, i) => {{
    b.classList.toggle('active', pages[i] === page);
  }});
  document.querySelectorAll('.page').forEach(p => p.classList.remove('active'));
  document.getElementById('page-' + page).classList.add('active');
  Object.keys(chartInstances).forEach(id => destroyChart(id));

  if (page === 'detalhe') {{
    document.getElementById('filtersBar').style.display = 'none';
    loadDetalhe();
    return;
  }}
  document.getElementById('filtersBar').style.display = 'flex';
  populateFilters();
  renderCurrentPage();
}}

// Init
populateFilters();
renderCurrentPage();
</script>
</body>
</html>"""

with open(OUT_PATH, 'w', encoding='utf-8') as f:
    f.write(HTML)

import os
size = os.path.getsize(OUT_PATH)
print(f"HTML gerado: {OUT_PATH}")
print(f"Tamanho: {size:,} bytes ({size/1024/1024:.1f} MB)")
has_fr = 'filteredRows' in HTML
print(f"Contém 'filteredRows': {has_fr}")
